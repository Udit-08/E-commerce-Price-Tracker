# import requests
# from bs4 import BeautifulSoup
# from selenium import webdriver
# from selenium.webdriver.common.by import By
# import pandas
# import streamlit
# from selenium.webdriver.common.keys import Keys
# from selenium.webdriver.chrome.service import Service
# from selenium.webdriver.support.ui import WebDriverWait
# from selenium.webdriver.support import expected_conditions as EC
# import time
# import schedule
# import openpyxl

# data = []


# service = Service(executable_path="D:\web_Scraping\chromedriver.exe")
# driver = webdriver.Chrome(service=service)

# driver.get("https://books.toscrape.com/catalogue/page-1.html")

# page_souce = driver.page_source

# soup = BeautifulSoup(page_souce,"html.parser")
# def get_info():
#     divs = soup.find_all("li",class_ = "col-xs-6 col-sm-4 col-md-3 col-lg-3")
#     for i in divs:
#         Title = i.find("a").find("img").get("alt").strip()
#         Stock = (i.find("p",class_ = "instock availability").text).strip()
#         Rating = i.find("p",class_ = "star-rating")["class"][1].strip()
#         Price = i.find("p",class_ = "price_color").text.strip()

#         data.append(
#             {
#                 "Title" : Title,
#                 "Stock" : Stock,
#                 "Price": Price,
#                 "Rating" : Rating
#             }
#         )       
#     driver.quit()

import requests
from bs4 import BeautifulSoup
from selenium import webdriver
from selenium.webdriver.chrome.service import Service
import openpyxl
from openpyxl.styles import Font, PatternFill

data = []

# ✅ Fix Windows path issue with raw string
service = Service(executable_path=r"D:\web_Scraping\chromedriver.exe")
driver = webdriver.Chrome(service=service)

driver.get("https://books.toscrape.com/catalogue/page-1.html")

page_source = driver.page_source
soup = BeautifulSoup(page_source, "html.parser")

def get_info():
    divs = soup.find_all("li", class_="col-xs-6 col-sm-4 col-md-3 col-lg-3")
    for i in divs:
        Title = i.find("a").find("img").get("alt").strip()
        Stock = (i.find("p", class_="instock availability").text).strip()
        Rating = i.find("p", class_="star-rating")["class"][1].strip()
        Price = i.find("p", class_="price_color").text.strip()

        # Remove currency symbol (£) and convert to float
        price_num = float(Price.lstrip("£"))

        data.append([Title, Stock, price_num, Rating])  

    driver.quit()

    # Create a new workbook
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Project 1"

    # Header row
    headers = ["Title", "Stock", "Price", "Rating"]
    ws.append(headers)

    # Style headers (bold)
    for cell in ws[1]:
        cell.font = Font(bold=True)

    # Write data rows
    for row in data:
        ws.append(row)

    # Format Price column (C column) as currency
    for cell in ws["C"][1:]:  # skip header
        cell.number_format = '£#,##0.00'

    # Highlight "Out of stock" rows in red
    red_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
    for row in ws.iter_rows(min_row=2, max_col=4):  # skip header
        stock_cell = row[1]  # column B = Stock
        if "out of stock" in stock_cell.value.lower():
            for cell in row:
                cell.fill = red_fill

    # Auto-adjust column widths
    for col in ws.columns:
        max_length = 0
        col_letter = col[0].column_letter
        for cell in col:
            try:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
            except:
                pass
        ws.column_dimensions[col_letter].width = max_length + 2

    # Save file
    wb.save("products_report.xlsx")

# Run function
get_info()
